import asyncio
import json
import os
import hashlib
import secrets
import time
import aiofiles
import psutil
from datetime import datetime, timedelta
from zoneinfo import ZoneInfo
from urllib.parse import quote
from collections import deque, defaultdict
from pathlib import Path
import socket
import base64
from io import BytesIO
import httpx
import re
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment

from fastapi import FastAPI, Request, HTTPException, Depends, WebSocket, WebSocketDisconnect
from fastapi.responses import Response, HTMLResponse, JSONResponse, RedirectResponse, StreamingResponse
from fastapi.middleware.cors import CORSMiddleware
import uvicorn
import logging

# ─── تنظیمات ──────────────────────────────────────────────────────────────────
logging.basicConfig(level=logging.INFO, format="%(asctime)s [%(levelname)s] %(message)s")
logger = logging.getLogger("Eagle-Gateway")

IRAN_TZ = ZoneInfo("Asia/Tehran")

CONFIG = {
    "port": int(os.environ.get("PORT", 8000)),
    "secret": os.environ.get("SECRET_KEY", secrets.token_urlsafe(32)),
    "host": os.environ.get("RAILWAY_PUBLIC_DOMAIN", os.environ.get("RENDER_EXTERNAL_URL", "localhost")),
    "admin_password": os.environ.get("ADMIN_PASSWORD", "123456"),
    "admin_username": os.environ.get("ADMIN_USERNAME", "admin"),
}

app = FastAPI(title="🪐 Eagle Gateway v10 Pro", docs_url=None, redoc_url=None)
app.add_middleware(CORSMiddleware, allow_origins=["*"], allow_credentials=True, allow_methods=["*"], allow_headers=["*"])

# ─── State ────────────────────────────────────────────────────────────────────
DATA_DIR = Path(os.environ.get("DATA_DIR", "/data"))
DATA_FILE = DATA_DIR / "eagle_state.json"
SAVE_LOCK = asyncio.Lock()

# ─── Workspace System ─────────────────────────────────────────────────────────
WORKSPACES: dict = {}
WORKSPACES_LOCK = asyncio.Lock()

# Super Admin (مدیر کل)
SUPER_ADMIN = {
    "username": os.environ.get("SUPER_ADMIN_USERNAME", "superadmin"),
    "password_hash": hashlib.sha256(f"{os.environ.get('SUPER_ADMIN_PASSWORD', 'superadmin123')}{CONFIG['secret']}".encode()).hexdigest(),
}

# ─── Auth ──────────────────────────────────────────────────────────────────────
SESSION_COOKIE = "eagle_session"
SESSION_TTL = 60 * 60 * 24 * 7
AUTH = {
    "password_hash": hashlib.sha256(f"{CONFIG['admin_password']}{CONFIG['secret']}".encode()).hexdigest(),
    "username": CONFIG["admin_username"],
}
SESSIONS: dict = {}
SESSIONS_LOCK = asyncio.Lock()

# ─── Default Settings ─────────────────────────────────────────────────────────
DEFAULT_SETTINGS = {
    "rgb_mode": False,
    "default_protocol": "vless-ws",
    "default_port": 443,
    "inbound_port": 443,
    "language": "fa",
    "theme": "dark",
    "telegram_bot_token": "",
    "telegram_chat_id": "",
    "clean_ips": [
        {"ip": "104.16.5.10", "provider": "MCI", "active": True},
        {"ip": "104.17.20.30", "provider": "Irancell", "active": True},
        {"ip": "104.18.10.20", "provider": "Hamrahe Aval", "active": True},
    ],
    "ad_filter_enabled": True,
    "porn_filter_enabled": True,
    "malware_filter_enabled": True,
    "backup_interval": 6,
    "multi_server": False,
    "servers": [],
    "routes": [],
    "log_retention_days": 30,
}

SETTINGS: dict = DEFAULT_SETTINGS.copy()

# ─── In-Memory State ─────────────────────────────────────────────────────────
connections: dict = {}
stats = {"total_bytes": 0, "total_requests": 0, "total_errors": 0, "start_time": time.time()}
error_logs: deque = deque(maxlen=50)
activity_logs: deque = deque(maxlen=200)
hourly_traffic: dict = defaultdict(int)
daily_traffic: dict = defaultdict(int)
device_connections: dict = {}
DEVICE_CONNECTIONS_LOCK = asyncio.Lock()
http_client: httpx.AsyncClient | None = None

PROTOCOLS = ("vless-ws", "xhttp-packet-up", "xhttp-stream-up", "xhttp-stream-one")
DEFAULT_PROTOCOL = "vless-ws"
DEFAULT_PORTS = [443, 8443, 2053, 2096, 2087, 2083, 8080]

FINGERPRINTS = {
    "chrome": "🌐 Chrome",
    "firefox": "🦊 Firefox",
    "safari": "🧭 Safari",
    "edge": "🌊 Edge",
    "ios": "📱 iOS",
    "android": "🤖 Android",
    "safari_ios": "🍏 Safari iOS",
    "random": "🎲 Random",
    "none": "🚫 None",
}

# ─────────────────────────────────────────────────────────────────────────────
# ===== Functions =====
# ─────────────────────────────────────────────────────────────────────────────

def now_ir() -> datetime:
    return datetime.now(IRAN_TZ)

def hash_password(pw: str) -> str:
    return hashlib.sha256(f"{pw}{CONFIG['secret']}".encode()).hexdigest()

def generate_uuid() -> str:
    h = secrets.token_hex(16)
    return f"{h[:8]}-{h[8:12]}-{h[12:16]}-{h[16:20]}-{h[20:32]}"

def get_host() -> str:
    return os.environ.get("RAILWAY_PUBLIC_DOMAIN", os.environ.get("RENDER_EXTERNAL_URL", CONFIG["host"]))

def fmt_bytes(b: int) -> str:
    if not b or b == 0:
        return "0 B"
    if b < 1024:
        return f"{b} B"
    if b < 1024**2:
        return f"{b/1024:.1f} KB"
    if b < 1024**3:
        return f"{b/1024**2:.2f} MB"
    if b < 1024**4:
        return f"{b/1024**3:.2f} GB"
    return f"{b/1024**4:.2f} TB"

def client_ip(request: Request) -> str:
    fwd = request.headers.get("x-forwarded-for")
    return fwd.split(",")[0].strip() if fwd else request.client.host if request.client else "نامشخص"

def uptime() -> str:
    secs = int(time.time() - stats["start_time"])
    return f"{secs//3600:02d}:{(secs%3600)//60:02d}:{secs%60:02d}"

def parse_size_to_bytes(value: float, unit: str) -> int:
    unit = unit.upper()
    if unit == "GB":
        return int(value * 1024 ** 3)
    if unit == "MB":
        return int(value * 1024 ** 2)
    if unit == "KB":
        return int(value * 1024)
    return int(value)

def is_link_expired(link: dict) -> bool:
    exp = link.get("expires_at")
    if not exp:
        return False
    try:
        return datetime.now() > datetime.fromisoformat(exp)
    except:
        return False

def is_link_allowed(link: dict | None) -> bool:
    if link is None:
        return False
    if not link.get("active", True):
        return False
    if is_link_expired(link):
        return False
    lb = link.get("limit_bytes", 0)
    if lb > 0 and link.get("used_bytes", 0) >= lb:
        return False
    return True

def generate_vless_link(uuid: str, host: str, remark: str = "", protocol: str = DEFAULT_PROTOCOL,
                        fingerprint: str = "chrome", port: int = 443, sni: str = None,
                        clean_ip: str = None) -> str:
    if not remark:
        remark = "عقاب"
    sni = sni or host
    actual_host = clean_ip if clean_ip else host
    
    if protocol == "vless-ws":
        path = f"/ws/{uuid}"
        params = {"encryption": "none", "security": "tls", "type": "ws", "host": host,
                  "path": path, "sni": sni, "fp": fingerprint, "alpn": "h2,http/1.1"}
    else:
        mode = protocol.replace("xhttp-", "")
        path = f"/xhttp-siz10/{mode}/{uuid}"
        params = {"encryption": "none", "security": "tls", "type": "xhttp", "mode": mode,
                  "host": host, "path": path, "sni": sni, "fp": fingerprint, "alpn": "h2,http/1.1"}
    query = "&".join(f"{k}={quote(str(v))}" for k, v in params.items())
    return f"vless://{uuid}@{actual_host}:{port}?{query}#{quote(remark)}"

def log_activity(kind: str, message: str, level: str = "info", workspace_id: str = None):
    activity_logs.append({
        "kind": kind,
        "level": level,
        "message": message,
        "time": datetime.now().isoformat(),
        "workspace_id": workspace_id
    })

async def remove_device_connection(uuid: str, client_ip: str):
    async with DEVICE_CONNECTIONS_LOCK:
        if uuid in device_connections and client_ip in device_connections[uuid]:
            device_connections[uuid].remove(client_ip)
            if not device_connections[uuid]:
                del device_connections[uuid]

# ─── Workspace Functions ──────────────────────────────────────────────────────

async def load_workspaces():
    """بارگذاری Workspace ها"""
    global WORKSPACES, SUPER_ADMIN, SETTINGS
    try:
        DATA_DIR.mkdir(parents=True, exist_ok=True)
        if DATA_FILE.exists():
            async with aiofiles.open(DATA_FILE, "r") as f:
                data = json.loads(await f.read())
                WORKSPACES = data.get("workspaces", {})
                if "super_admin" in data:
                    SUPER_ADMIN["username"] = data["super_admin"].get("username", SUPER_ADMIN["username"])
                    SUPER_ADMIN["password_hash"] = data["super_admin"].get("password_hash", SUPER_ADMIN["password_hash"])
                if "settings" in data:
                    SETTINGS.update(data["settings"])
                logger.info(f"📂 Loaded {len(WORKSPACES)} workspaces")
    except Exception as e:
        logger.warning(f"Load workspaces error: {e}")

async def save_workspaces():
    """ذخیره Workspace ها"""
    async with SAVE_LOCK:
        try:
            data = {
                "workspaces": WORKSPACES,
                "super_admin": {
                    "username": SUPER_ADMIN["username"],
                    "password_hash": SUPER_ADMIN["password_hash"],
                },
                "settings": SETTINGS,
                "saved_at": datetime.now().isoformat()
            }
            async with aiofiles.open(DATA_FILE, "w") as f:
                await f.write(json.dumps(data, ensure_ascii=False, indent=2))
        except Exception as e:
            logger.warning(f"Save workspaces error: {e}")

async def get_workspace_data(workspace_id: str) -> dict:
    """دریافت دیتای Workspace با fallback"""
    ws = WORKSPACES.get(workspace_id, {})
    if not ws:
        return {
            "links": {},
            "settings": DEFAULT_SETTINGS.copy(),
            "daily_traffic": {},
            "activity_logs": [],
            "stats": {"total_bytes": 0, "total_requests": 0, "total_errors": 0},
            "quota_limit_bytes": 0,
            "quota_used_bytes": 0,
            "quota_reset_date": None,
            "is_quota_exceeded": False,
        }
    return ws

async def check_workspace_quota(workspace_id: str) -> tuple[bool, str]:
    """بررسی سهمیه Workspace"""
    ws = WORKSPACES.get(workspace_id)
    if not ws:
        return False, "Workspace یافت نشد"
    
    limit = ws.get("quota_limit_bytes", 0)
    used = ws.get("quota_used_bytes", 0)
    
    # اگر سهمیه 0 باشه = نامحدود
    if limit == 0:
        return True, "نامحدود"
    
    # اگر مصرف از سهمیه بیشتر شده = غیرفعال
    if used >= limit:
        ws["is_quota_exceeded"] = True
        await save_workspaces()
        return False, f"سهمیه تمام شده! {fmt_bytes(used)} / {fmt_bytes(limit)}"
    
    return True, f"{fmt_bytes(used)} / {fmt_bytes(limit)}"

async def update_workspace_usage(workspace_id: str, bytes_used: int):
    """به‌روزرسانی مصرف Workspace"""
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            return
        ws = WORKSPACES[workspace_id]
        ws["quota_used_bytes"] = ws.get("quota_used_bytes", 0) + bytes_used
        # بررسی دوباره سهمیه
        limit = ws.get("quota_limit_bytes", 0)
        if limit > 0 and ws["quota_used_bytes"] >= limit:
            ws["is_quota_exceeded"] = True
        await save_workspaces()

# ─── Session Functions ──────────────────────────────────────────────────────

async def create_session() -> str:
    token = secrets.token_urlsafe(32)
    async with SESSIONS_LOCK:
        SESSIONS[token] = time.time() + SESSION_TTL
    return token

async def is_valid_session(token: str | None) -> bool:
    if not token:
        return False
    async with SESSIONS_LOCK:
        exp = SESSIONS.get(token)
        if exp is None or exp < time.time():
            SESSIONS.pop(token, None)
            return False
        return True

async def destroy_session(token: str | None):
    if token:
        async with SESSIONS_LOCK:
            SESSIONS.pop(token, None)

# ─── Auth Dependencies ──────────────────────────────────────────────────────

async def require_super_admin(req: Request):
    """فقط Super Admin"""
    token = req.cookies.get(SESSION_COOKIE)
    if not await is_valid_session(token):
        raise HTTPException(401, "unauthorized")
    session_data = SESSIONS.get(token, {})
    if session_data.get("user_type") != "super_admin":
        raise HTTPException(403, "دسترسی محدود به Super Admin")
    return token

async def require_workspace_admin(req: Request):
    """ادمین یک Workspace"""
    token = req.cookies.get(SESSION_COOKIE)
    if not await is_valid_session(token):
        raise HTTPException(401, "unauthorized")
    session_data = SESSIONS.get(token, {})
    if session_data.get("user_type") != "workspace_admin":
        raise HTTPException(403, "دسترسی محدود به ادمین Workspace")
    workspace_id = session_data.get("workspace_id")
    if not workspace_id or workspace_id not in WORKSPACES:
        raise HTTPException(404, "Workspace یافت نشد")
    return workspace_id

# ─── Telegram Functions ──────────────────────────────────────────────────────

async def tg_send(chat_id, text, kb=None):
    token = SETTINGS.get("telegram_bot_token", "")
    if not token:
        return False
    url = f"https://api.telegram.org/bot{token}/sendMessage"
    payload = {"chat_id": chat_id, "text": text, "parse_mode": "HTML", "disable_web_page_preview": True}
    if kb:
        payload["reply_markup"] = json.dumps({"inline_keyboard": kb})
    try:
        async with httpx.AsyncClient(timeout=3.0) as client:
            r = await client.post(url, json=payload)
            return r.status_code == 200
    except:
        return False

async def tg_doc(chat_id, content, filename):
    token = SETTINGS.get("telegram_bot_token", "")
    if not token:
        return False
    url = f"https://api.telegram.org/bot{token}/sendDocument"
    try:
        async with httpx.AsyncClient(timeout=10.0) as client:
            r = await client.post(url, data={"chat_id": chat_id}, files={"document": (filename, BytesIO(content.encode()), "text/plain")})
            return r.status_code == 200
    except:
        return False

async def setup_webhook():
    token = SETTINGS.get("telegram_bot_token", "")
    if not token:
        return
    host = get_host()
    try:
        async with httpx.AsyncClient(timeout=5.0) as client:
            await client.post(f"https://api.telegram.org/bot{token}/setWebhook", json={"url": f"https://{host}/webhook/telegram"})
            logger.info("✅ Webhook set")
    except:
        pass

# ─── Startup ─────────────────────────────────────────────────────────────────

@app.on_event("startup")
async def startup():
    global http_client
    http_client = httpx.AsyncClient(timeout=10.0)
    await load_workspaces()
    await setup_webhook()
    asyncio.create_task(auto_backup())
    logger.info(f"🪐 Eagle Gateway started on port {CONFIG['port']}")

@app.on_event("shutdown")
async def shutdown():
    await save_workspaces()
    if http_client:
        await http_client.aclose()

# ─── Auto Backup ─────────────────────────────────────────────────────────────

async def auto_backup():
    while True:
        try:
            await asyncio.sleep(SETTINGS.get("backup_interval", 6) * 3600)
            await save_workspaces()
            chat_id = SETTINGS.get("telegram_chat_id")
            token = SETTINGS.get("telegram_bot_token")
            if chat_id and token:
                async with aiofiles.open(DATA_FILE, "r") as f:
                    content = await f.read()
                await tg_doc(chat_id, content, f"backup_{now_ir().strftime('%Y-%m-%d_%H-%M')}.json")
            logger.info("✅ Auto backup completed")
        except Exception as e:
            logger.error(f"Auto backup error: {e}")
            await asyncio.sleep(300)

# ─── ===== API: Login ===== ──────────────────────────────────────────────────

@app.post("/api/login")
async def api_login(req: Request):
    body = await req.json()
    username = body.get("username", "").strip()
    password = body.get("password", "").strip()
    workspace_id = body.get("workspace_id", "").strip()
    
    # ===== Super Admin =====
    if username == SUPER_ADMIN["username"] and hash_password(password) == SUPER_ADMIN["password_hash"]:
        token = await create_session()
        async with SESSIONS_LOCK:
            SESSIONS[token] = {"user_type": "super_admin", "workspace_id": None}
        resp = JSONResponse({"ok": True, "user_type": "super_admin", "is_super_admin": True})
        resp.set_cookie(SESSION_COOKIE, token, max_age=SESSION_TTL if body.get("remember") else None, httponly=True, samesite="lax", path="/")
        return resp
    
    # ===== Workspace Admin =====
    if workspace_id and workspace_id in WORKSPACES:
        ws = WORKSPACES[workspace_id]
        if ws.get("admin_username") == username and hash_password(password) == ws.get("admin_password_hash"):
            # بررسی سهمیه
            limit = ws.get("quota_limit_bytes", 0)
            used = ws.get("quota_used_bytes", 0)
            if limit > 0 and used >= limit:
                return JSONResponse({"ok": False, "error": "سهمیه این Workspace تمام شده است!"}, status_code=403)
            
            token = await create_session()
            async with SESSIONS_LOCK:
                SESSIONS[token] = {"user_type": "workspace_admin", "workspace_id": workspace_id}
            resp = JSONResponse({"ok": True, "user_type": "workspace_admin", "workspace_name": ws.get("name")})
            resp.set_cookie(SESSION_COOKIE, token, max_age=SESSION_TTL if body.get("remember") else None, httponly=True, samesite="lax", path="/")
            resp.set_cookie("eagle_workspace", workspace_id, max_age=SESSION_TTL if body.get("remember") else None, httponly=True, samesite="lax", path="/")
            return resp
    
    raise HTTPException(401, "نام کاربری یا رمز عبور اشتباه است")

@app.post("/api/logout")
async def api_logout(req: Request):
    await destroy_session(req.cookies.get(SESSION_COOKIE))
    resp = JSONResponse({"ok": True})
    resp.delete_cookie(SESSION_COOKIE, path="/")
    resp.delete_cookie("eagle_workspace", path="/")
    return resp

@app.get("/api/me")
async def api_me(req: Request):
    token = req.cookies.get(SESSION_COOKIE)
    is_auth = await is_valid_session(token)
    session_data = SESSIONS.get(token, {}) if token else {}
    user_type = session_data.get("user_type")
    workspace_id = session_data.get("workspace_id")
    
    return {
        "authenticated": is_auth,
        "user_type": user_type,
        "workspace_id": workspace_id,
        "is_super_admin": user_type == "super_admin",
    }

# ─── ===== API: Workspace Management (Super Admin) ===== ────────────────────

@app.get("/api/admin/workspaces")
async def admin_list_workspaces(_=Depends(require_super_admin)):
    """لیست همه Workspace ها"""
    result = []
    for wid, ws in WORKSPACES.items():
        limit = ws.get("quota_limit_bytes", 0)
        used = ws.get("quota_used_bytes", 0)
        result.append({
            "id": wid,
            "name": ws.get("name"),
            "admin_username": ws.get("admin_username"),
            "created_at": ws.get("created_at"),
            "links_count": len(ws.get("links", {})),
            "active": ws.get("active", True),
            "quota_limit_gb": round(limit / (1024**3), 2) if limit > 0 else 0,
            "quota_used_gb": round(used / (1024**3), 2),
            "quota_percent": round((used / limit) * 100, 1) if limit > 0 else 0,
            "is_quota_exceeded": ws.get("is_quota_exceeded", False),
            "quota_reset_date": ws.get("quota_reset_date"),
        })
    return {"workspaces": result}

@app.post("/api/admin/workspaces")
async def admin_create_workspace(req: Request, _=Depends(require_super_admin)):
    """ساخت Workspace جدید"""
    body = await req.json()
    name = body.get("name", "").strip()
    admin_username = body.get("admin_username", "").strip()
    admin_password = body.get("admin_password", "").strip()
    quota_gb = float(body.get("quota_gb", 0))
    
    if not name or not admin_username or not admin_password:
        raise HTTPException(400, "نام، نام کاربری و رمز عبور الزامی است")
    
    if len(admin_password) < 4:
        raise HTTPException(400, "رمز عبور حداقل 4 کاراکتر")
    
    # بررسی تکراری نبودن
    for wid, ws in WORKSPACES.items():
        if ws.get("admin_username") == admin_username:
            raise HTTPException(400, "این نام کاربری قبلاً استفاده شده است")
    
    workspace_id = f"ws_{secrets.token_hex(8)}"
    async with WORKSPACES_LOCK:
        WORKSPACES[workspace_id] = {
            "name": name,
            "admin_username": admin_username,
            "admin_password_hash": hash_password(admin_password),
            "created_at": datetime.now().isoformat(),
            "active": True,
            "quota_limit_bytes": int(quota_gb * 1024**3) if quota_gb > 0 else 0,
            "quota_used_bytes": 0,
            "quota_reset_date": None,
            "is_quota_exceeded": False,
            "settings": DEFAULT_SETTINGS.copy(),
            "links": {},
            "daily_traffic": {},
            "activity_logs": [],
            "stats": {"total_bytes": 0, "total_requests": 0, "total_errors": 0},
        }
    
    await save_workspaces()
    log_activity("workspace", f"Workspace {name} ساخته شد", "ok")
    return {"ok": True, "workspace_id": workspace_id}

@app.delete("/api/admin/workspaces/{workspace_id}")
async def admin_delete_workspace(workspace_id: str, _=Depends(require_super_admin)):
    """حذف Workspace"""
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        name = WORKSPACES[workspace_id].get("name", "بدون نام")
        del WORKSPACES[workspace_id]
    
    await save_workspaces()
    log_activity("workspace", f"Workspace {name} حذف شد", "err")
    return {"ok": True}

@app.post("/api/admin/workspaces/{workspace_id}/quota")
async def admin_set_quota(workspace_id: str, req: Request, _=Depends(require_super_admin)):
    """تنظیم سهمیه Workspace"""
    body = await req.json()
    quota_gb = float(body.get("quota_gb", 0))
    reset_date = body.get("reset_date")
    
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        ws = WORKSPACES[workspace_id]
        ws["quota_limit_bytes"] = int(quota_gb * 1024**3) if quota_gb > 0 else 0
        if reset_date:
            ws["quota_reset_date"] = reset_date
        ws["is_quota_exceeded"] = False
    
    await save_workspaces()
    log_activity("quota", f"سهمیه {WORKSPACES[workspace_id].get('name')} به {quota_gb}GB تغییر کرد", "ok")
    return {"ok": True}

@app.post("/api/admin/workspaces/{workspace_id}/quota/reset")
async def admin_reset_quota(workspace_id: str, _=Depends(require_super_admin)):
    """ریست مصرف Workspace"""
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        ws = WORKSPACES[workspace_id]
        ws["quota_used_bytes"] = 0
        ws["is_quota_exceeded"] = False
        # ریست مصرف کاربران هم
        for uid, link in ws.get("links", {}).items():
            link["used_bytes"] = 0
    
    await save_workspaces()
    log_activity("quota", f"مصرف {WORKSPACES[workspace_id].get('name')} ریست شد", "ok")
    return {"ok": True}

# ─── ===== API: Workspace (برای ادمین Workspace) ===== ──────────────────────

@app.get("/api/workspace/info")
async def workspace_get_info(workspace_id=Depends(require_workspace_admin)):
    """دریافت اطلاعات Workspace جاری"""
    ws = WORKSPACES.get(workspace_id, {})
    limit = ws.get("quota_limit_bytes", 0)
    used = ws.get("quota_used_bytes", 0)
    return {
        "id": workspace_id,
        "name": ws.get("name"),
        "created_at": ws.get("created_at"),
        "links_count": len(ws.get("links", {})),
        "settings": ws.get("settings", {}),
        "quota": {
            "limit_gb": round(limit / (1024**3), 2) if limit > 0 else 0,
            "used_gb": round(used / (1024**3), 2),
            "percent": round((used / limit) * 100, 1) if limit > 0 else 0,
            "is_exceeded": ws.get("is_quota_exceeded", False),
            "reset_date": ws.get("quota_reset_date"),
        }
    }

# ─── ===== API: Workspace Links ===== ────────────────────────────────────────

@app.get("/api/workspace/links")
async def workspace_list_links(req: Request, workspace_id=Depends(require_workspace_admin)):
    """لیست کاربران در Workspace جاری"""
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    host = get_host()
    
    result = []
    for uid, d in links.items():
        ports = d.get("ports", [443])
        first_port = ports[0] if ports else 443
        active = d.get("active", True) and not is_link_expired(d)
        label = d.get("label", "کاربر")
        clean_ip = d.get("clean_ip")
        
        result.append({
            "uuid": uid,
            **d,
            "ports": ports,
            "expired": is_link_expired(d),
            "has_password": d.get("password_hash") is not None,
            "vless_link": generate_vless_link(uid, host, remark=f"عقاب-{label}",
                protocol=d.get("protocol", DEFAULT_PROTOCOL),
                fingerprint=d.get("fingerprint", "chrome"),
                port=first_port, clean_ip=clean_ip),
            "sub_url": f"https://{host}/sub/{uid}"
        })
    
    result.sort(key=lambda x: x["created_at"], reverse=True)
    return {"links": result}

@app.post("/api/workspace/links")
async def workspace_create_link(req: Request, workspace_id=Depends(require_workspace_admin)):
    """ساخت کاربر جدید در Workspace جاری"""
    # بررسی سهمیه Workspace
    can_use, msg = await check_workspace_quota(workspace_id)
    if not can_use:
        raise HTTPException(403, f"سهمیه Workspace تمام شده! {msg}")
    
    body = await req.json()
    label = (body.get("label") or "کاربر").strip()[:60]
    quota = float(body.get("limit_value") or 0)
    limit_bytes = 0 if quota <= 0 else parse_size_to_bytes(quota, "GB")
    exp_days = int(body.get("expires_days") or 30)
    expires_at = (datetime.now() + timedelta(days=exp_days)).isoformat() if exp_days > 0 else None
    protocol = body.get("protocol") or DEFAULT_PROTOCOL
    if protocol not in PROTOCOLS:
        protocol = DEFAULT_PROTOCOL
    max_devices = int(body.get("max_devices", 0))
    fingerprint = body.get("fingerprint", "chrome")
    if fingerprint not in FINGERPRINTS:
        fingerprint = "chrome"
    password_hash = hash_password(body.get("password", "").strip()) if body.get("password") else None
    clean_ip = body.get("clean_ip") or None
    ports = body.get("ports", [443])
    if not isinstance(ports, list) or not ports:
        ports = [443]
    ports = [p for p in ports if isinstance(p, int) and 1 <= p <= 65535] or [443]

    uid = generate_uuid()
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        WORKSPACES[workspace_id]["links"][uid] = {
            "label": label,
            "limit_bytes": limit_bytes,
            "used_bytes": 0,
            "created_at": datetime.now().isoformat(),
            "active": True,
            "expires_at": expires_at,
            "note": "",
            "is_default": False,
            "sub_id": None,
            "protocol": protocol,
            "max_devices": max_devices,
            "fingerprint": fingerprint,
            "password_hash": password_hash,
            "ports": ports,
            "clean_ip": clean_ip,
        }
    
    await save_workspaces()
    log_activity("link", f"کانفیگ «{label}» در {WORKSPACES[workspace_id].get('name')} ساخته شد", "ok", workspace_id)
    
    host = get_host()
    main_link = generate_vless_link(uid, host, remark=f"عقاب-{label}", protocol=protocol,
                                    fingerprint=fingerprint, port=ports[0], clean_ip=clean_ip)
    return {"uuid": uid, **WORKSPACES[workspace_id]["links"][uid], "has_password": password_hash is not None,
            "vless_link": main_link, "sub_url": f"https://{host}/sub/{uid}"}

@app.patch("/api/workspace/links/{uid}")
async def workspace_update_link(uid: str, req: Request, workspace_id=Depends(require_workspace_admin)):
    """ویرایش کاربر در Workspace جاری"""
    body = await req.json()
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        ws = WORKSPACES[workspace_id]
        if uid not in ws.get("links", {}):
            raise HTTPException(404, "کاربر یافت نشد")
        link = ws["links"][uid]
        
        if link.get("password_hash"):
            pw = body.get("password", "").strip()
            if not pw or hash_password(pw) != link["password_hash"]:
                raise HTTPException(403, "رمز کانفیگ اشتباه است")
        if "label" in body:
            link["label"] = str(body["label"])[:60]
        if "limit_value" in body:
            lv = float(body["limit_value"] or 0)
            link["limit_bytes"] = 0 if lv <= 0 else parse_size_to_bytes(lv, "GB")
        if "expires_days" in body:
            ed = int(body["expires_days"] or 0)
            link["expires_at"] = (datetime.now() + timedelta(days=ed)).isoformat() if ed > 0 else None
        if "max_devices" in body:
            link["max_devices"] = int(body["max_devices"])
        if "fingerprint" in body and body["fingerprint"] in FINGERPRINTS:
            link["fingerprint"] = body["fingerprint"]
        if "clean_ip" in body:
            link["clean_ip"] = body["clean_ip"] or None
        if "active" in body:
            link["active"] = bool(body["active"])
        if "reset_usage" in body and body["reset_usage"]:
            link["used_bytes"] = 0
        if "ports" in body and isinstance(body["ports"], list):
            ports = [p for p in body["ports"] if isinstance(p, int) and 1 <= p <= 65535]
            if ports:
                link["ports"] = ports
    
    await save_workspaces()
    return {"ok": True}

@app.delete("/api/workspace/links/{uid}")
async def workspace_delete_link(uid: str, req: Request, workspace_id=Depends(require_workspace_admin)):
    """حذف کاربر از Workspace جاری"""
    body = await req.json()
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        ws = WORKSPACES[workspace_id]
        if uid not in ws.get("links", {}):
            raise HTTPException(404, "کاربر یافت نشد")
        link = ws["links"][uid]
        if link.get("password_hash"):
            pw = body.get("password", "").strip()
            if not pw or hash_password(pw) != link["password_hash"]:
                raise HTTPException(403, "رمز کانفیگ اشتباه است")
        label = link.get("label", uid)
        del ws["links"][uid]
    
    await save_workspaces()
    log_activity("link", f"کانفیگ «{label}» از {WORKSPACES[workspace_id].get('name')} حذف شد", "err", workspace_id)
    return {"ok": True}

# ─── ===== API: Workspace Settings ===== ──────────────────────────────────────

@app.get("/api/workspace/settings")
async def workspace_get_settings(workspace_id=Depends(require_workspace_admin)):
    """دریافت تنظیمات Workspace جاری"""
    ws = WORKSPACES.get(workspace_id, {})
    return ws.get("settings", DEFAULT_SETTINGS.copy())

@app.post("/api/workspace/settings")
async def workspace_update_settings(req: Request, workspace_id=Depends(require_workspace_admin)):
    """به‌روزرسانی تنظیمات Workspace جاری"""
    body = await req.json()
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        for key, value in body.items():
            if key in DEFAULT_SETTINGS:
                WORKSPACES[workspace_id]["settings"][key] = value
    
    await save_workspaces()
    return {"ok": True}

# ─── ===== API: Workspace Stats ===== ────────────────────────────────────────

@app.get("/api/workspace/stats")
async def workspace_stats(workspace_id=Depends(require_workspace_admin)):
    """آمار Workspace جاری"""
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    stats_data = ws.get("stats", {})
    daily = ws.get("daily_traffic", {})
    today = now_ir().strftime("%Y-%m-%d")
    
    active_links = sum(1 for l in links.values() if l.get("active", True) and not is_link_expired(l))
    total_used = sum(l.get("used_bytes", 0) for l in links.values())
    today_traffic = daily.get(today, 0)
    
    return {
        "links_count": len(links),
        "active_links": active_links,
        "total_used": total_used,
        "total_used_fmt": fmt_bytes(total_used),
        "today_traffic": today_traffic,
        "today_traffic_fmt": fmt_bytes(today_traffic),
        "total_requests": stats_data.get("total_requests", 0),
        "total_errors": stats_data.get("total_errors", 0),
        "connections": len(connections),
    }

# ─── ===== API: Workspace Clean IP ===== ─────────────────────────────────────

@app.get("/api/workspace/clean-ips")
async def workspace_get_clean_ips(workspace_id=Depends(require_workspace_admin)):
    """دریافت لیست IP‌های تمیز Workspace"""
    ws = WORKSPACES.get(workspace_id, {})
    return {"clean_ips": ws.get("settings", {}).get("clean_ips", [])}

@app.post("/api/workspace/clean-ips")
async def workspace_add_clean_ip(req: Request, workspace_id=Depends(require_workspace_admin)):
    """افزودن IP تمیز به Workspace"""
    body = await req.json()
    ip = body.get("ip", "").strip()
    provider = body.get("provider", "Unknown")
    
    if not ip:
        raise HTTPException(400, "IP نمی‌تواند خالی باشد")
    
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        clean_ips = WORKSPACES[workspace_id]["settings"].get("clean_ips", [])
        for item in clean_ips:
            if item.get("ip") == ip:
                raise HTTPException(400, "این IP قبلاً اضافه شده است")
        clean_ips.append({"ip": ip, "provider": provider, "active": True})
        WORKSPACES[workspace_id]["settings"]["clean_ips"] = clean_ips
    
    await save_workspaces()
    return {"ok": True}

@app.delete("/api/workspace/clean-ips/{ip}")
async def workspace_delete_clean_ip(ip: str, workspace_id=Depends(require_workspace_admin)):
    """حذف IP تمیز از Workspace"""
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        clean_ips = WORKSPACES[workspace_id]["settings"].get("clean_ips", [])
        WORKSPACES[workspace_id]["settings"]["clean_ips"] = [item for item in clean_ips if item.get("ip") != ip]
    
    await save_workspaces()
    return {"ok": True}

# ─── ===== API: Workspace Filters ===== ─────────────────────────────────────

@app.get("/api/workspace/filters")
async def workspace_get_filters(workspace_id=Depends(require_workspace_admin)):
    """دریافت وضعیت فیلترها"""
    ws = WORKSPACES.get(workspace_id, {})
    settings = ws.get("settings", {})
    return {
        "ad_filter_enabled": settings.get("ad_filter_enabled", True),
        "porn_filter_enabled": settings.get("porn_filter_enabled", True),
        "malware_filter_enabled": settings.get("malware_filter_enabled", True),
    }

@app.post("/api/workspace/filters")
async def workspace_update_filters(req: Request, workspace_id=Depends(require_workspace_admin)):
    """به‌روزرسانی وضعیت فیلترها"""
    body = await req.json()
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        settings = WORKSPACES[workspace_id]["settings"]
        if "ad_filter_enabled" in body:
            settings["ad_filter_enabled"] = bool(body["ad_filter_enabled"])
        if "porn_filter_enabled" in body:
            settings["porn_filter_enabled"] = bool(body["porn_filter_enabled"])
        if "malware_filter_enabled" in body:
            settings["malware_filter_enabled"] = bool(body["malware_filter_enabled"])
    
    await save_workspaces()
    return {"ok": True}

# ─── ===== API: Workspace Routes ===== ───────────────────────────────────────

@app.get("/api/workspace/routes")
async def workspace_get_routes(workspace_id=Depends(require_workspace_admin)):
    """دریافت لیست مسیرها"""
    ws = WORKSPACES.get(workspace_id, {})
    return {"routes": ws.get("settings", {}).get("routes", [])}

@app.post("/api/workspace/routes")
async def workspace_add_route(req: Request, workspace_id=Depends(require_workspace_admin)):
    """افزودن مسیر جدید"""
    body = await req.json()
    domain = body.get("domain", "").strip()
    route_type = body.get("type", "direct")
    target = body.get("target", "")
    
    if not domain:
        raise HTTPException(400, "دامنه نمی‌تواند خالی باشد")
    
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        routes = WORKSPACES[workspace_id]["settings"].get("routes", [])
        routes.append({"domain": domain, "type": route_type, "target": target, "active": True})
        WORKSPACES[workspace_id]["settings"]["routes"] = routes
    
    await save_workspaces()
    return {"ok": True}

@app.delete("/api/workspace/routes/{domain}")
async def workspace_delete_route(domain: str, workspace_id=Depends(require_workspace_admin)):
    """حذف مسیر"""
    async with WORKSPACES_LOCK:
        if workspace_id not in WORKSPACES:
            raise HTTPException(404, "Workspace یافت نشد")
        routes = WORKSPACES[workspace_id]["settings"].get("routes", [])
        WORKSPACES[workspace_id]["settings"]["routes"] = [r for r in routes if r.get("domain") != domain]
    
    await save_workspaces()
    return {"ok": True}

# ─── ===== API: Workspace Export ===== ───────────────────────────────────────

@app.get("/api/workspace/export/excel")
async def workspace_export_excel(workspace_id=Depends(require_workspace_admin)):
    """خروجی Excel از کاربران Workspace"""
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    
    wb = Workbook()
    ws_excel = wb.active
    ws_excel.title = "کاربران"
    
    headers = ["نام", "UUID", "مصرف", "سهمیه", "وضعیت", "انقضا", "دستگاه‌ها", "فینگرپرینت"]
    ws_excel.append(headers)
    
    header_fill = PatternFill(start_color="7C6BFF", end_color="7C6BFF", fill_type="solid")
    header_font = Font(color="FFFFFF", bold=True)
    for col, _ in enumerate(headers, 1):
        cell = ws_excel.cell(row=1, column=col)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
    
    for uid, link in links.items():
        used = link.get("used_bytes", 0)
        limit = link.get("limit_bytes", 0)
        active = link.get("active", True) and not is_link_expired(link)
        status = "فعال" if active else ("منقضی" if is_link_expired(link) else "غیرفعال")
        expires_at = link.get("expires_at", "نامحدود")
        if expires_at and expires_at != "نامحدود":
            try:
                expires_at = datetime.fromisoformat(expires_at).strftime("%Y-%m-%d")
            except:
                pass
        
        ws_excel.append([
            link.get("label", "بدون نام"),
            uid[:8] + "..." + uid[-8:],
            fmt_bytes(used),
            fmt_bytes(limit) if limit > 0 else "نامحدود",
            status,
            expires_at,
            link.get("max_devices", 0),
            link.get("fingerprint", "chrome"),
        ])
    
    for col in range(1, len(headers) + 1):
        ws_excel.column_dimensions[chr(64 + col)].width = 20
    
    output = BytesIO()
    wb.save(output)
    output.seek(0)
    
    return StreamingResponse(
        output,
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=users_{now_ir().strftime('%Y-%m-%d')}.xlsx"}
    )

# ─── ===== API: Workspace Activity Logs ===== ────────────────────────────────

@app.get("/api/workspace/logs")
async def workspace_get_logs(req: Request, workspace_id=Depends(require_workspace_admin)):
    """دریافت لاگ‌های Workspace"""
    ws = WORKSPACES.get(workspace_id, {})
    logs = ws.get("activity_logs", [])
    limit = int(req.query_params.get("limit", 50))
    level = req.query_params.get("level", "")
    search = req.query_params.get("search", "").lower()
    
    result = list(logs)[-limit:]
    
    if level:
        result = [l for l in result if l.get("level") == level]
    if search:
        result = [l for l in result if search in l.get("message", "").lower()]
    
    return {"logs": result}

# ─── ===== WebSocket Tunnel ===== ─────────────────────────────────────────────

RELAY_BUF = 512 * 1024

def _ws_ip(ws):
    fwd = ws.headers.get("x-forwarded-for")
    return fwd.split(",")[0].strip() if fwd else ws.client.host if ws.client else "نامشخص"

async def check_device_limit(uuid: str, client_ip: str, workspace_id: str = None) -> bool:
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    link = links.get(uuid)
    if not link or link.get("max_devices", 0) == 0:
        return True
    max_devices = link["max_devices"]
    async with DEVICE_CONNECTIONS_LOCK:
        current = device_connections.get(uuid, [])
        if client_ip in current:
            return True
        if len(current) >= max_devices:
            return False
        device_connections.setdefault(uuid, []).append(client_ip)
        return True

async def parse_vless_header(chunk: bytes):
    if len(chunk) < 24:
        raise ValueError("chunk too small")
    pos = 1 + 16
    addon_len = chunk[pos]
    pos += 1 + addon_len
    command = chunk[pos]
    pos += 1
    port = int.from_bytes(chunk[pos:pos+2], "big")
    pos += 2
    addr_type = chunk[pos]
    pos += 1
    if addr_type == 1:
        address = ".".join(str(b) for b in chunk[pos:pos+4])
        pos += 4
    elif addr_type == 2:
        dlen = chunk[pos]
        pos += 1
        address = chunk[pos:pos+dlen].decode("utf-8", errors="ignore")
        pos += dlen
    elif addr_type == 3:
        ab = chunk[pos:pos+16]
        pos += 16
        address = ":".join(f"{ab[i]:02x}{ab[i+1]:02x}" for i in range(0, 16, 2))
    else:
        raise ValueError(f"unknown addr type: {addr_type}")
    return command, address, port, chunk[pos:]

async def check_and_use(uid: str, n: int, workspace_id: str) -> bool:
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    link = links.get(uid)
    if not link or not is_link_allowed(link):
        return False
    
    # بروزرسانی مصرف کاربر
    link["used_bytes"] = link.get("used_bytes", 0) + n
    
    # بروزرسانی مصرف Workspace
    ws["quota_used_bytes"] = ws.get("quota_used_bytes", 0) + n
    
    # بروزرسانی آمار روزانه
    today = now_ir().strftime("%Y-%m-%d")
    daily = ws.get("daily_traffic", {})
    daily[today] = daily.get(today, 0) + n
    ws["daily_traffic"] = daily
    
    # بررسی سهمیه Workspace
    limit = ws.get("quota_limit_bytes", 0)
    used = ws.get("quota_used_bytes", 0)
    if limit > 0 and used >= limit:
        ws["is_quota_exceeded"] = True
        await save_workspaces()
        return False
    
    # هشدار 80%
    if limit > 0 and used / limit > 0.8 and not link.get("alert_80"):
        link["alert_80"] = True
        log_activity("warning", f"⚠️ مصرف {link.get('label')} به 80% رسید", "warn", workspace_id)
        chat_id = SETTINGS.get("telegram_chat_id")
        if chat_id:
            await tg_send(chat_id, f"⚠️ هشدار مصرف!\nکاربر: {link.get('label')}\nمصرف: {fmt_bytes(used)} / {fmt_bytes(limit)}")
    
    return True

async def relay_ws_to_tcp(ws, writer, conn_id, uid, workspace_id):
    try:
        while True:
            msg = await ws.receive()
            if msg["type"] == "websocket.disconnect":
                break
            data = msg.get("bytes") or (msg.get("text") or "").encode()
            if not data:
                continue
            if not await check_and_use(uid, len(data), workspace_id):
                await ws.close(code=1008, reason="quota/disabled")
                break
            if conn_id in connections:
                connections[conn_id]["bytes"] = connections[conn_id].get("bytes", 0) + len(data)
            writer.write(data)
            if writer.transport.get_write_buffer_size() > RELAY_BUF:
                await writer.drain()
    except:
        pass
    finally:
        try:
            writer.write_eof()
        except:
            pass

async def relay_tcp_to_ws(ws, reader, conn_id, uid, workspace_id):
    first = True
    try:
        while True:
            data = await reader.read(RELAY_BUF)
            if not data:
                break
            if not await check_and_use(uid, len(data), workspace_id):
                await ws.close(code=1008, reason="quota/disabled")
                break
            if conn_id in connections:
                connections[conn_id]["bytes"] = connections[conn_id].get("bytes", 0) + len(data)
            payload = (b"\x00\x00" + data) if first else data
            first = False
            await ws.send_bytes(payload)
    except:
        pass

@app.websocket("/ws/{uuid}")
async def websocket_tunnel(ws: WebSocket, uuid: str):
    await ws.accept()
    
    # تشخیص Workspace از هدر یا کوکی
    workspace_id = ws.headers.get("X-Workspace-ID")
    if not workspace_id:
        cookie_header = ws.headers.get("cookie", "")
        for cookie in cookie_header.split(";"):
            if "eagle_workspace=" in cookie:
                workspace_id = cookie.split("eagle_workspace=")[1].strip()
                break
    
    if not workspace_id or workspace_id not in WORKSPACES:
        await ws.close(code=1008, reason="workspace not found")
        return
    
    client_ip = _ws_ip(ws)
    ws_data = WORKSPACES.get(workspace_id, {})
    links = ws_data.get("links", {})
    link = links.get(uuid)
    
    if not link:
        await ws.close(code=1008, reason="user not found")
        return
    
    if not is_link_allowed(link):
        await ws.close(code=1008, reason="not authorized")
        return
    
    # بررسی سهمیه Workspace
    can_use, msg = await check_workspace_quota(workspace_id)
    if not can_use:
        await ws.close(code=1008, reason=f"quota exceeded: {msg}")
        return
    
    if link.get("max_devices", 0) > 0 and not await check_device_limit(uuid, client_ip, workspace_id):
        await ws.close(code=1008, reason="device limit exceeded")
        return

    conn_id = secrets.token_urlsafe(6)
    connections[conn_id] = {"uuid": uuid, "ip": client_ip, "transport": "vless-ws",
                            "connected_at": datetime.now().isoformat(), "bytes": 0}
    writer = None

    try:
        first_msg = await asyncio.wait_for(ws.receive(), timeout=15.0)
        if first_msg["type"] == "websocket.disconnect":
            return
        first_chunk = first_msg.get("bytes") or (first_msg.get("text") or "").encode()
        if not first_chunk:
            return
        command, address, port, payload = await parse_vless_header(first_chunk)
        
        if not await check_and_use(uuid, len(first_chunk), workspace_id):
            await ws.close(code=1008, reason="quota/disabled")
            return
        
        if conn_id in connections:
            connections[conn_id]["bytes"] = connections[conn_id].get("bytes", 0) + len(first_chunk)

        reader, writer = await asyncio.wait_for(asyncio.open_connection(address, port), timeout=10.0)
        if writer:
            sock = writer.transport.get_extra_info('socket')
            if sock:
                sock.setsockopt(socket.IPPROTO_TCP, socket.TCP_NODELAY, 1)
        if payload:
            writer.write(payload)
            await writer.drain()

        done, pending = await asyncio.wait(
            {asyncio.create_task(relay_ws_to_tcp(ws, writer, conn_id, uuid, workspace_id)),
             asyncio.create_task(relay_tcp_to_ws(ws, reader, conn_id, uuid, workspace_id))},
            return_when=asyncio.FIRST_COMPLETED
        )
        for t in pending:
            t.cancel()
            try:
                await t
            except:
                pass
        await save_workspaces()
    except:
        pass
    finally:
        if writer:
            try:
                writer.close()
                await writer.wait_closed()
            except:
                pass
        connections.pop(conn_id, None)
        await remove_device_connection(uuid, client_ip)

# ─── ===== Subscriptions ===== ────────────────────────────────────────────────

@app.get("/sub/{uuid}")
async def subscription_single(req: Request, uuid: str):
    import base64
    
    # تشخیص Workspace از هدر یا کوکی
    workspace_id = req.headers.get("X-Workspace-ID")
    if not workspace_id:
        cookie_header = req.headers.get("cookie", "")
        for cookie in cookie_header.split(";"):
            if "eagle_workspace=" in cookie:
                workspace_id = cookie.split("eagle_workspace=")[1].strip()
                break
    
    if not workspace_id or workspace_id not in WORKSPACES:
        return HTMLResponse("""<html><body><h1>Workspace یافت نشد</h1></body></html>""", status_code=404)
    
    ua = req.headers.get("user-agent", "").lower()
    is_browser = any(b in ua for b in ["chrome", "firefox", "safari", "edge", "opera", "brave", "msie", "trident"])
    
    ws = WORKSPACES.get(workspace_id, {})
    links = ws.get("links", {})
    link = links.get(uuid)
    
    if not link:
        if is_browser:
            return HTMLResponse("""<html><body><h1>کاربر یافت نشد</h1></body></html>""", status_code=404)
        raise HTTPException(404, "user not found")
    
    if not is_link_allowed(link):
        if is_browser:
            return HTMLResponse("""<html><body><h1>کاربر غیرفعال</h1></body></html>""", status_code=403)
        raise HTTPException(403, "user disabled or expired")
    
    host = get_host()
    ports = link.get("ports", [443])
    clean_ip = link.get("clean_ip")
    vless_links = []
    
    for port in ports:
        remark = f"{link.get('label','کاربر')}-p{port}" if len(ports) > 1 else link.get('label', 'کاربر')
        vless_links.append(generate_vless_link(uuid, host, remark=remark,
                                               protocol=link.get("protocol", DEFAULT_PROTOCOL),
                                               fingerprint=link.get("fingerprint", "chrome"),
                                               port=port, clean_ip=clean_ip))
    
    if not is_browser:
        userinfo = f"upload=0&download={link.get('used_bytes',0)}&total={link.get('limit_bytes',0)}"
        try:
            if link.get("expires_at"):
                exp_ts = int(datetime.fromisoformat(link["expires_at"].replace('Z', '+00:00')).timestamp())
                userinfo += f"&expire={exp_ts}"
        except:
            pass
        content = base64.b64encode("\n".join(vless_links).encode()).decode()
        return Response(content=content, media_type="text/plain",
                        headers={"Subscription-Userinfo": userinfo,
                                 "profile-title": quote(link.get("label", "کاربر")),
                                 "profile-update-interval": "12"})
    
    from pages import get_sub_page_html
    active_conns = [c for c in connections.values() if c.get("uuid") == uuid]
    link_data = {**link, "expired": is_link_expired(link), "active_connections": len(active_conns),
                 "active_connections_list": active_conns, "vless_links": vless_links,
                 "vless_link": vless_links[0] if vless_links else "", "sub_url": f"https://{host}/sub/{uuid}"}
    return HTMLResponse(content=get_sub_page_html(uuid, link_data))

# ─── ===== HTML Pages ===== ──────────────────────────────────────────────────

from pages import LOGIN_HTML, DASHBOARD_HTML

@app.get("/login")
async def login_page(req: Request):
    if await is_valid_session(req.cookies.get(SESSION_COOKIE)):
        return RedirectResponse(url="/dashboard")
    return HTMLResponse(content=LOGIN_HTML)

@app.get("/dashboard")
async def dashboard(req: Request):
    token = req.cookies.get(SESSION_COOKIE)
    if not await is_valid_session(token):
        return RedirectResponse(url="/login")
    
    session_data = SESSIONS.get(token, {})
    if session_data.get("user_type") == "super_admin":
        # Super Admin به صفحه مدیریت Workspace ها میره
        return HTMLResponse(content=DASHBOARD_HTML)
    elif session_data.get("user_type") == "workspace_admin":
        return HTMLResponse(content=DASHBOARD_HTML)
    
    return RedirectResponse(url="/login")

@app.get("/")
async def root():
    return HTMLResponse("""<!DOCTYPE html><html><head><meta charset="UTF-8"><title>🪐 Eagle Gateway</title>
    <style>body{font-family:sans-serif;background:#0a0a0f;color:#fff;display:flex;align-items:center;justify-content:center;height:100vh;margin:0}
    .card{text-align:center;padding:40px;background:rgba(20,20,40,0.7);border-radius:20px;border:1px solid rgba(100,80,255,0.2)}
    h1{font-size:48px;margin:0}.sub{color:#888}a{color:#7C6BFF;text-decoration:none}</style></head>
    <body><div class="card"><h1>🪐</h1><h2>Eagle Gateway v10 Pro</h2><p class="sub">پنل مدیریت فیلترشکن</p><a href="/login">ورود →</a></div></body></html>""")

if __name__ == "__main__":
    uvicorn.run("main:app", host="0.0.0.0", port=CONFIG["port"], log_level="info", workers=1)
